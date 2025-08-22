import os
import re
import streamlit as st
import tempfile
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter

# === Data fields ===
data_fields = [
    ('Blocks', r'Number of blocks\s*:\s*(\d+)'),
    ('Trips', r'Number of in-service trips\s*:\s*(\d+)'),
    ('Off Service Duration', r'Off-service duration\s*:\s*([\dhm]+)'),
    ('In Service Duration', r'In-service duration\s*:\s*([\dhm]+)'),
    ('Loading Duration', r'Loading duration\s*:\s*([\dhm]+)'),
    ('Layover Duration', r'Layover duration\s*:\s*([\dhm]+)'),
    ('Platform Hours', r'Total duration\s*:\s*([\dhm]+)'),
    ('In-service Distance (KM)', r'In-service distance\s*:\s*([\d\.]+)'),
    ('Revenue Hours', None),  # Computed
]

# Desired order
file_order = ["WDY Stats.prt", "SAT Stats.prt", "SUN Stats.prt"]

# Excel colors
color_map = {
    'WDY Stats.prt': '0070C0',  # blue
    'SAT Stats.prt': 'FF0000',  # red
    'SUN Stats.prt': '000000',  # black
}
beige_fill = PatternFill(start_color='F5F5DC', end_color='F5F5DC', fill_type='solid')
revenue_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')  # Yellow

# === Utils ===
def parse_time_string(time_str):
    match = re.match(r'(\d+)h(\d+)', time_str)
    if match:
        hours = int(match.group(1))
        minutes = int(match.group(2))
        return round(hours + minutes / 60, 2)
    return 0.0

def extract_booking_code(lines):
    for line in lines:
        match = re.search(r'Booking:\s*(\d+)', line)
        if match:
            return match.group(1)
    return "unknown"

def extract_data(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    booking_code = extract_booking_code(lines)
    data = {}
    current_route = None

    for line in lines:
        route_match = re.match(r'\s*Route\s+(\S+)\s+(.*)', line)
        if route_match:
            route_num = route_match.group(1)
            route_name = route_match.group(2).strip()
            current_route = f"{route_num} {route_name}"
            data[current_route] = {field[0]: 0 if field[0] != 'Revenue Hours' else 0.0 for field in data_fields}
            continue

        if current_route:
            for field_name, regex in data_fields:
                if regex is None:
                    continue
                m = re.search(regex, line)
                if m:
                    val = m.group(1).strip()
                    if 'Duration' in field_name or field_name == 'Platform Hours':
                        val = parse_time_string(val)
                    else:
                        try:
                            val = float(val) if '.' in val else int(val)
                        except:
                            val = 0
                    data[current_route][field_name] = val

    # Compute Revenue Hours
    for route in data:
        in_service = data[route].get('In Service Duration', 0)
        layover = data[route].get('Layover Duration', 0)
        revenue_hours = round(float(in_service) + float(layover), 2)
        data[route]['Revenue Hours'] = revenue_hours

        if route.strip().lower() == "65 senior shopper":
            for key in data[route]:
                data[route][key] = round(data[route][key] / 2, 2)

    return booking_code, data

def route_sort_key(route):
    match = re.match(r'(\d+)', route)
    return int(match.group(1)) if match else float('inf')

def write_to_excel(all_data, booking_code):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Vehicle Schedule Overview"

    # Build headers
    headers = ["Route number"]
    for f in file_order:
        if f in all_data:
            for field_name, _ in data_fields:
                headers.append(field_name)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = beige_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)

    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.cell(row=1, column=1, value="Route number").fill = beige_fill
    ws.cell(row=1, column=1).font = Font(bold=True)

    start_col = 2
    for f in file_order:
        if f in all_data:
            color = color_map.get(f, 'FFFFFF')
            ws.merge_cells(start_row=1, start_column=start_col,
                           end_row=1, end_column=start_col + len(data_fields) - 1)
            cell = ws.cell(row=1, column=start_col, value=f.split()[0])
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
            start_col += len(data_fields)

    # Gather and sort routes
    all_routes = sorted({r for d in all_data.values() for r in d.keys()}, key=route_sort_key)

    row_idx = 3
    for route in all_routes:
        ws.cell(row=row_idx, column=1, value=route)
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center", vertical="center")

        for file_idx, f in enumerate(file_order):
            if f not in all_data:
                continue
            data = all_data[f]
            col_offset = 2 + file_idx * len(data_fields)
            route_data = data.get(route, {})
            for field_idx, (field_name, _) in enumerate(data_fields):
                val = route_data.get(field_name, "")
                cell = ws.cell(row=row_idx, column=col_offset + field_idx, value=val)
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if field_name == "Revenue Hours":
                    cell.fill = revenue_fill
        row_idx += 1

    # === Remove TAXI row(s) ===
    for row in range(3, row_idx):
        val = ws.cell(row=row, column=1).value
        if val and "TAXI" in str(val).upper():
            ws.delete_rows(row)
            row_idx -= 1
            break  # remove first TAXI found

    # Totals row with formulas
    total_row_idx = row_idx
    ws.cell(row=total_row_idx, column=1, value="Total").font = Font(bold=True)
    ws.cell(row=total_row_idx, column=1).alignment = Alignment(horizontal="center", vertical="center")

    for file_idx, f in enumerate(file_order):
        if f not in all_data:
            continue
        col_offset = 2 + file_idx * len(data_fields)
        for field_idx, (field_name, _) in enumerate(data_fields):
            col_letter = get_column_letter(col_offset + field_idx)
            formula = f"=SUM({col_letter}3:{col_letter}{total_row_idx-1})"
            cell = ws.cell(row=total_row_idx, column=col_offset + field_idx, value=formula)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    safe_booking_code = booking_code if booking_code.lower() != "unknown" else "Unknown"

    tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    output_file = tmp_file.name
    wb.save(output_file)
    return output_file, f"{safe_booking_code} - Vehicle Schedule Overview.xlsx"


# === Streamlit UI ===
st.title("Bus Schedule Reporter")
st.subheader("Upload PRT files")

uploaded_files = st.file_uploader("Upload PRT files", type=["prt"], accept_multiple_files=True)

if uploaded_files:
    all_data = {}
    booking_code = "unknown"

    with tempfile.TemporaryDirectory() as tmpdirname:
        for uploaded_file in uploaded_files:
            file_path = os.path.join(tmpdirname, uploaded_file.name)
            with open(file_path, "wb") as f:
                f.write(uploaded_file.read())

            bc, data = extract_data(file_path)
            all_data[uploaded_file.name] = data
            if bc.lower() != "unknown" and booking_code == "unknown":
                booking_code = bc

        excel_data, filename = write_to_excel(all_data, booking_code)

        with open(excel_data, "rb") as f:
            st.download_button(
                label="Download Excel Report",
                data=f,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
